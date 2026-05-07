import win32com.client
import os
import sys
import openpyxl
import re

def convert_xlsb_to_xlsx(xlsb_path, xlsx_path):
    print(f"Conversione di {xlsb_path} in {xlsx_path}...")
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)
    
    # Crea la directory .tmp se non esiste
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    
    try:
        # Usa UpdateLinks=0 per ignorare aggiornamenti link, CorruptLoad=1 per forzare apertura
        wb = excel.Workbooks.Open(xlsb_path, ReadOnly=True, UpdateLinks=0, CorruptLoad=1)
        wb.SaveAs(xlsx_path, FileFormat=51) # 51 = xlOpenXMLWorkbook
        wb.Close(False)
        print("Conversione completata.")
    except Exception as e:
        print(f"Errore durante la conversione COM: {e}")
        sys.exit(1)
    finally:
        excel.Quit()

def analyze_formulas(xlsx_path):
    print("\nAnalisi delle formule...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n=========================================")
        print(f"FOGLIO: {sheet_name}")
        print(f"=========================================")
        
        # Intestazioni (Riga 1)
        headers = []
        for col_idx in range(1, min(sheet.max_column + 1, 30)):
            val = sheet.cell(row=1, column=col_idx).value
            if val is not None:
                headers.append(f"Col {col_idx}: {val}")
        if headers:
            print("Intestazioni rilevate (riga 1):")
            for h in headers:
                print(f"  - {h}")
        else:
            print("Nessuna intestazione chiara nella prima riga.")
            
        formulas = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    f_str = str(cell.value)
                    
                    # Rimuoviamo i numeri di riga per trovare pattern ripetuti
                    # Es: =SUM(A1:A10) diventa =SUM(AN:AN)
                    f_pattern = re.sub(r'\d+', 'N', f_str)
                    
                    if f_pattern not in formulas:
                        formulas[f_pattern] = {
                            'example_formula': f_str,
                            'example_cell': cell.coordinate,
                            'count': 1
                        }
                    else:
                        formulas[f_pattern]['count'] += 1
                        
        if not formulas:
            print("Nessuna formula trovata in questo foglio.")
        else:
            print(f"\nTrovati {len(formulas)} pattern di formule unici:\n")
            # Ordiniamo per frequenza
            sorted_formulas = sorted(formulas.items(), key=lambda item: item[1]['count'], reverse=True)
            for pattern, data in sorted_formulas:
                print(f"  -> Cella Esempio: {data['example_cell']}")
                print(f"  -> Formula Esempio: {data['example_formula']}")
                print(f"  -> Pattern Base: {pattern}")
                print(f"  -> Ripetuta: {data['count']} volte")
                print("-" * 60)

if __name__ == '__main__':
    xlsb_file = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\virgi\Desktop\LAVORO\File Contabile Anatema.xlsb"
    xlsx_file = os.path.join(r"C:\Users\virgi\Desktop\PAOLO\gestionale-ordini\.tmp", "File_Contabile_Temp.xlsx")
    
    convert_xlsb_to_xlsx(xlsb_file, xlsx_file)
    analyze_formulas(xlsx_file)
